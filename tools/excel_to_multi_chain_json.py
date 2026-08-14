#!/usr/bin/env python3
"""
excel_to_multi_chain_json.py
============================
提示词审阅表 Excel → 多链生产 JSON

读取用户审阅后的提示词审阅表 Excel，按主角色分组生成多链预编码 JSON。
同一图片只 Load 一次，连到多个 H3EncodeConditioning（合并 LoadImage）。

用法:
    python excel_to_multi_chain_json.py <审阅表.xlsx> <输出目录> [--by-shot]

参数:
    审阅表.xlsx   — shot_md_to_excel.py 生成、用户审阅后的 Excel
    输出目录       — 多链 JSON 输出路径（如 preencode_by_char/）
    --by-shot      — 可选，按镜头顺序而非按角色分组（每镜一个JSON）
"""

import json
import sys
import os
import re
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


def sanitize_filename(name):
    """清理文件名中的非法字符，截断过长名称。

    Windows 非法字符: \\ / : * ? " < > |
    同时去除 Excel 单元格中可能混入的 Markdown 格式残留（**、| 等）。
    """
    # 去掉 Markdown 粗体标记和管道符分隔的多余内容
    name = re.sub(r'\*\*', '', name)
    # 如果名称中包含 |，只取第一部分（通常是角色名）
    if '|' in name:
        name = name.split('|')[0].strip()
    # 替换 Windows 非法字符为下划线
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    # 去除首尾空格和点
    name = name.strip().strip('.')
    # 截断过长名称
    if len(name) > 50:
        name = name[:50]
    return name if name else "unnamed"


# ── 常量 ─────────────────────────────────────────────────────────────

# MiniMaxH3ReferenceToVideo 的 ref_image 槽位对应的 input 索引
SLOT_INPUT = {0: 3, 1: 4, 2: 5, 3: 13, 4: 14, 5: 15}

# 共享节点模板
CLIP_MODEL = "qwen3vl_32b_minimax_h3_int8_convrot.safetensors"
VIDEO_VAE = "minimax_h3_video_vae_fp16.safetensors"
AUDIO_VAE = "minimax_h3_audio_vae_fp32.safetensors"


# ── ID 管理器 ─────────────────────────────────────────────────────────

class IDGen:
    def __init__(self, start=1):
        self._next = start

    def node(self):
        n = self._next
        self._next += 1
        return n

    def link(self):
        n = self._next
        self._next += 1
        return n


# ── 节点工厂 ─────────────────────────────────────────────────────────

def make_node(nid, ntype, title, pos, size, widgets_values=None,
              inputs=None, outputs=None, color=None, bgcolor=None):
    node = {
        "id": nid,
        "type": ntype,
        "pos": list(pos),
        "size": list(size),
        "flags": {},
        "order": 0,
        "mode": 0,
        "inputs": inputs or [],
        "outputs": outputs or [],
        "properties": {"Node name for S&R": ntype},
        "widgets_values": widgets_values or [],
        "title": title,
    }
    if color:
        node["color"] = color
    if bgcolor:
        node["bgcolor"] = bgcolor
    return node


def make_output(name, dtype, links=None, slot=0):
    return {"name": name, "type": dtype, "links": links or [], "slot_index": slot}


def make_input(name, dtype, link=None, widget_name=None):
    inp = {"name": name, "type": dtype, "link": link}
    if widget_name:
        inp["widget"] = {"name": widget_name}
    return inp


# ── Excel 读取 ───────────────────────────────────────────────────────

def read_excel(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)

    # Sheet1: 提示词审阅
    ws1 = wb["提示词审阅"]
    headers = [cell.value for cell in ws1[1]]

    shots = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        shot = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                shot[h] = str(row[i]).strip() if row[i] else ""
        shots.append(shot)

    # Sheet2: 美术资产路径
    asset_paths = {}
    if "美术资产路径" in wb.sheetnames:
        ws2 = wb["美术资产路径"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            asset_type = str(row[0]).strip()
            asset_name = str(row[1]).strip()
            input_path = str(row[2]).strip() if row[2] else ""
            if input_path:
                asset_paths[(asset_type, asset_name)] = input_path

    return shots, asset_paths


# ── 提示词改写 ───────────────────────────────────────────────────────

def compute_slots(char, scene, prop, sup_chars):
    """计算槽位分配 {资产名: (ref_image槽位索引, 图片编号)}。

    有主角色: 主角色=ref0(图片1), 场景=ref1(图片2), 道具=ref2(图片3), 配角1=ref3(图片4)
    纯场景:   场景=ref0(图片1), 道具=ref1(图片2), 配角1=ref2(图片3)
    超出槽位的资产（道具2/配角2）不进槽位，仅保留正文描述。
    """
    slots = {}
    if char and char != "(纯场景)":
        slots[char] = (0, 1)
        if scene:
            slots[scene] = (1, 2)
        if prop:
            slots[prop] = (2, 3)
        if sup_chars:
            slots[sup_chars[0]] = (3, 4)
    else:
        next_slot = 0
        if scene:
            slots[scene] = (next_slot, next_slot + 1)
            next_slot += 1
        if prop:
            slots[prop] = (next_slot, next_slot + 1)
            next_slot += 1
        if sup_chars:
            slots[sup_chars[0]] = (next_slot, next_slot + 1)
            next_slot += 1
    return slots


def build_ref_sentences(slots, char, scene, prop, sup_chars):
    """构建前缀参考句（使用<图片N>，与正文一致）。"""
    refs = []
    if char and char != "(纯场景)":
        refs.append(f"使用<图片1>作为{char}的身份参考（驱动其身份）")
    if scene:
        refs.append(f"使用<图片{slots[scene][1]}>作为{scene}场景环境参考")
    if prop:
        refs.append(f"使用<图片{slots[prop][1]}>作为{prop}道具参考")
    if sup_chars:
        sup = sup_chars[0]
        refs.append(f"使用<图片{slots[sup][1]}>作为{sup}的身份参考（驱动其身份）")
    return "，".join(refs) + "。" if refs else ""


def build_ref_constraints(slots, char, scene, prop, sup_chars):
    """构建【参考图约束】节（按槽位顺序）。"""
    entries = []
    if char and char != "(纯场景)":
        entries.append(
            f"以 <图片1>（{char}角色定妆照） 为基准，严格保持{char}的外观、体态、服饰完全一致，不得身份漂移；"
            f"初始姿态从 <图片1>（{char}角色定妆照） 呈现的状态开始")
    if scene:
        n = slots[scene][1]
        entries.append(f"以 <图片{n}>（{scene}场景参考图） 为基准，保持场景空间布局、光线方向完全一致")
    if prop:
        n = slots[prop][1]
        entries.append(f"以 <图片{n}>（{prop}道具参考图） 为基准，保持{prop}外观与摆放一致")
    if sup_chars:
        sup = sup_chars[0]
        n = slots[sup][1]
        entries.append(
            f"以 <图片{n}>（{sup}角色定妆照） 为基准，严格保持{sup}的外观、体态、服饰完全一致，不得身份漂移；"
            f"初始姿态从 <图片{n}>（{sup}角色定妆照） 呈现的状态开始")
    return "；".join(entries) + "。" if entries else "无参考图（纯文本/黑场）"


def build_constraints(slots, char, scene, prop, sup_chars, duration):
    """构建【约束条件】节（按槽位顺序）。"""
    entries = []
    if char and char != "(纯场景)":
        entries.append("图片1中的角色外观严格与 <图片1> 一致")
    if scene:
        n = slots[scene][1]
        entries.append(f"画面场景与 <图片{n}> 一致")
    if prop:
        n = slots[prop][1]
        entries.append(f"{prop}与 <图片{n}> 一致")
    if sup_chars:
        sup = sup_chars[0]
        n = slots[sup][1]
        entries.append(f"图片{n}中的{sup}外观严格与 <图片{n}> 一致")
    dur_str = str(int(duration)) if float(duration).is_integer() else str(duration)
    entries.append(f"16:9，{dur_str}秒精确")
    return "；".join(entries) + "。"


def split_sections(prompt):
    """按【节名】分割提示词，返回 [(name, content), ...]。"""
    pattern = re.compile(r'【([^】]+)】')
    matches = list(pattern.finditer(prompt))
    sections = []
    if not matches:
        return [("", prompt)]
    if matches[0].start() > 0:
        sections.append(("", prompt[:matches[0].start()]))
    for i, m in enumerate(matches):
        name = m.group(1)
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(prompt)
        sections.append((name, prompt[start:end]))
    return sections


def replace_names(text, name_map):
    """替换角色名，长名优先避免重叠（占位符法）。"""
    if not name_map:
        return text
    placeholders = {}
    for i, old in enumerate(sorted(name_map.keys(), key=len, reverse=True)):
        ph = f"\x00PH{i}\x00"
        text = text.replace(old, ph)
        placeholders[ph] = name_map[old]
    for ph, new in placeholders.items():
        text = text.replace(ph, new)
    return text


def rewrite_prompt(prompt, char, scene, prop, sup_chars, duration):
    """重建提示词：统一参考句 + 参考图约束 + 约束条件，使图片编号与槽位分配一致。"""
    slots = compute_slots(char, scene, prop, sup_chars)

    # 角色名替换映射（仅正文节；参考图约束/约束条件单独重建）
    name_map = {}
    if char and char != "(纯场景)":
        name_map[char] = "图片1中的角色"
    if sup_chars:
        sup = sup_chars[0]
        if sup in slots:
            name_map[sup] = f"图片{slots[sup][1]}中的{sup}"
    # 配角2无槽位，不替换（保留原名）

    sections = split_sections(prompt)
    rebuilt = []
    for name, content in sections:
        if name == "参考图约束":
            rebuilt.append(f"【参考图约束】{build_ref_constraints(slots, char, scene, prop, sup_chars)}")
        elif name == "约束条件":
            rebuilt.append(f"【约束条件】{build_constraints(slots, char, scene, prop, sup_chars, duration)}")
        elif name == "":
            rebuilt.append(content)
        else:
            rebuilt.append(f"【{name}】{replace_names(content, name_map)}")

    body = "".join(rebuilt)
    ref_str = build_ref_sentences(slots, char, scene, prop, sup_chars)
    return ref_str + body


# ── 多链 JSON 生成 ───────────────────────────────────────────────────

def build_shared_nodes(idgen, resolution=0.5):
    """创建循环体外共享节点，返回节点列表和输出引用。

    拆解模式（H3EncodeConditioning）：
    - 预编码只做 CLIP 编码，不需要 Video VAE / Audio VAE / ResolutionSelector
    - VAE 编码与分辨率由生成阶段 H3ReencodeFromCache 处理
    """
    nodes = []

    # CLIPLoader
    clip_id = idgen.node()
    nodes.append(make_node(
        clip_id, "CLIPLoader", "H3 CLIP (Qwen3VL 32B)",
        [-1800, -400], [300, 82],
        widgets_values=[CLIP_MODEL, "minimax", "default"],
        outputs=[make_output("CLIP", "CLIP", links=[], slot=0)],
        color="#322", bgcolor="#533",
    ))

    refs = {
        "clip_id": clip_id, "clip_out": 0,
    }
    return nodes, refs


def _clean_name(name):
    """从复合名称中提取纯名称，去除非法文件名字符。
    
    Excel 中的资产名可能包含 | 分隔的附加信息（如场景参考、道具参考），
    以及 Markdown 粗体标记 **。本函数提取纯名称并替换非法字符。
    """
    if not name:
        return name
    name = re.sub(r'\*\*', '', name)
    if '|' in name:
        name = name.split('|')[0].strip()
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    return name.strip()


_PLACEHOLDER_FILE = "h3_placeholder.png"


def _candidate_input_dirs():
    """候选 ComfyUI input 目录列表（脚本独立运行，无法依赖 folder_paths）。"""
    dirs = []
    env = os.environ.get("COMFYUI_INPUT_DIR")
    if env:
        dirs.append(env)
    # 常见安装路径
    dirs.extend([
        r"F:\02aidraw\ComfyUI-aki-v3\ComfyUI\input",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "input"),
    ])
    return [d for d in dirs if d and os.path.isdir(d)]


def _ensure_placeholder_image():
    """确保各候选 input 目录存在占位图 h3_placeholder.png（纯灰，标准库生成）。

    场景/道具资产缺失时，LoadImage 挂这个占位图，避免 ref 槽位缺失，
    用户在 ComfyUI 中看到灰色占位图后手动替换为真实资产图。
    """
    import struct
    import zlib

    def _png_bytes(size=256):
        w = h = size
        row = b"\x00" + b"\x80\x80\x80\xff" * w
        raw = row * h

        def chunk(typ, data):
            return (struct.pack(">I", len(data)) + typ + data
                    + struct.pack(">I", zlib.crc32(typ + data) & 0xFFFFFFFF))

        ihdr = struct.pack(">IIBBBBB", w, h, 8, 6, 0, 0, 0)
        return (b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr)
                + chunk(b"IDAT", zlib.compress(raw)) + chunk(b"IEND", b""))

    created = []
    for d in _candidate_input_dirs():
        target = os.path.join(d, _PLACEHOLDER_FILE)
        if not os.path.isfile(target):
            try:
                with open(target, "wb") as f:
                    f.write(_png_bytes())
                created.append(target)
            except OSError:
                pass
    return _PLACEHOLDER_FILE, created


def build_loadimage(idgen, asset_name, image_path, x, y):
    """创建一个 LoadImage 节点。"""
    nid = idgen.node()
    node = make_node(
        nid, "LoadImage", f"Load {asset_name}",
        [x, y], [300, 82],
        widgets_values=[image_path, "image"],
        outputs=[
            make_output("IMAGE", "IMAGE", links=[], slot=0),
            make_output("MASK", "MASK", links=[], slot=1),
        ],
        color="#322", bgcolor="#533",
    )
    return nid, node


def build_loadaudio(idgen, asset_name, audio_path, x, y):
    """创建一个 LoadAudio 节点。"""
    nid = idgen.node()
    node = make_node(
        nid, "LoadAudio", f"Load Audio ({asset_name})",
        [x, y], [300, 82],
        widgets_values=[audio_path],
        outputs=[make_output("AUDIO", "AUDIO", links=[], slot=0)],
        color="#223", bgcolor="#335",
    )
    return nid, node


def build_primitive_float(idgen, duration, x, y):
    """创建 PrimitiveFloat 节点（时长）。"""
    nid = idgen.node()
    node = make_node(
        nid, "PrimitiveFloat", f"Duration ({duration}s)",
        [x, y], [300, 82],
        widgets_values=[float(duration)],
        outputs=[make_output("FLOAT", "FLOAT", links=[], slot=0)],
        color="#322", bgcolor="#533",
    )
    return nid, node


def build_comfy_math(idgen, float_id, float_out, x, y):
    """创建 ComfyMathExpression 节点（帧数计算）。"""
    nid = idgen.node()
    link_id = idgen.link()
    node = make_node(
        nid, "ComfyMathExpression", "Frame Count",
        [x, y], [300, 82],
        inputs=[make_input("values.a", "FLOAT", link=link_id)],
        widgets_values=["max(5, round(a * 24)) + (5 - (max(5, round(a * 24)) % 17)) % 17"],
        outputs=[
            make_output("FLOAT", "FLOAT", links=[], slot=0),
            make_output("INT", "INT", links=[], slot=1),
            make_output("BOOL", "BOOL", links=[], slot=2),
        ],
        color="#322", bgcolor="#533",
    )
    return nid, node, link_id


def _extract_number(val, default):
    """从字符串中提取首个数字，用于时长/分辨率等含单位的值（如 '11秒' → 11.0）。"""
    if val is None:
        return default
    m = re.search(r"-?\d+(?:\.\d+)?", str(val))
    if m:
        return float(m.group(0))
    return default


def build_shot_chain(idgen, shot, shared_refs, load_nodes, asset_paths, x_offset, y_offset):
    """为单个镜头创建一条链：PrimitiveStringMultiline → H3EncodeConditioning → H3SaveConditioning。"""
    nodes = []
    links = []

    shot_id = shot.get("镜头编号", "unknown")
    duration = _extract_number(shot.get("时长", "5"), 5.0)

    # 获取参演资产（清理复合名称，与 generate_group_json 保持一致）
    char = _clean_name(shot.get("参演角色1", ""))
    scene = _clean_name(shot.get("场景设置1", "") or shot.get("场景设置2", ""))
    prop = _clean_name(shot.get("参演道具1", ""))
    sup_chars = []
    for i in [2, 3]:
        sc = _clean_name(shot.get(f"参演角色{i}", ""))
        if sc:
            sup_chars.append(sc)

    # ── Per-shot PrimitiveFloat (duration) ──
    dur_id = idgen.node()
    dur_node = make_node(
        dur_id, "PrimitiveFloat", f"Duration ({shot_id}, {duration}s)",
        [x_offset - 600, y_offset], [300, 82],
        widgets_values=[float(duration)],
        outputs=[make_output("FLOAT", "FLOAT", links=[], slot=0)],
        color="#322", bgcolor="#533",
    )
    nodes.append(dur_node)

    # ── PrimitiveStringMultiline (Text Multiline) ──
    # 注意：不能用 easy promptLine！它会按 \n 把九分节提示词拆成多行 STRING，
    # ComfyUI 对 list 输入做展开（_async_map_node_over_list），导致
    # H3EncodeConditioning / H3SaveConditioning 各执行 N 次（行数=执行次数），
    # 一次运行生成 N 个 .pt。PrimitiveStringMultiline 输出完整字符串不拆分。
    pl_id = idgen.node()
    raw_prompt = shot.get("完整提示词", "")
    rewritten = rewrite_prompt(raw_prompt, char, scene, prop, sup_chars, duration)

    pl_node = make_node(
        pl_id, "PrimitiveStringMultiline", f"H3 Prompt ({shot_id})",
        [x_offset - 300, y_offset + 520], [300, 82],
        inputs=[
            make_input("value", "STRING", link=None, widget_name="value"),
        ],
        outputs=[
            make_output("STRING", "STRING", links=[], slot=0),
        ],
        widgets_values=[rewritten],
    )
    nodes.append(pl_node)

    # ── H3EncodeConditioning (CLIP-only 预编码，无 VAE、无分辨率) ──
    enc_id = idgen.node()

    # 构建 ref_image 输入（与 H3SaveConditioning 共享同一 LoadImage 源）
    # 槽位分配（与 compute_slots 一致）：
    #   有主角色: 主=ref0(图片1)/场景=ref1(图片2)/道具=ref2(图片3)/配角1=ref3(图片4)
    #   纯场景:   场景=ref0(图片1)/道具=ref1(图片2)/配角1=ref2(图片3)
    slots = compute_slots(char, scene, prop, sup_chars)
    asset_types = {}
    if char and char != "(纯场景)":
        asset_types[char] = "char"
    if scene:
        asset_types[scene] = "scene"
    if prop:
        asset_types[prop] = "prop"
    for sup in sup_chars:
        asset_types[sup] = "char"

    enc_inputs = []
    ref_image_links = {}  # index -> (load_node_id, link_id)，供 H3SaveConditioning 复用同一来源
    slot_assets = {si: name for name, (si, _pn) in slots.items()}
    for si in range(4):
        if si in slot_assets:
            name = slot_assets[si]
            atype = asset_types.get(name, "char")
            ln = load_nodes.get((atype, name))
            if atype == "scene" and ln is None:
                ln = load_nodes.get(("scene", "__placeholder__"))
            if atype == "prop" and ln is None:
                ln = load_nodes.get(("prop", "__placeholder__"))
            if ln:
                lid = idgen.link()
                enc_inputs.append(make_input(f"ref_image_{si}", "IMAGE", link=lid))
                links.append([lid, ln, 0, enc_id, 2 + si, "IMAGE"])
                ref_image_links[si] = (ln, lid)
            else:
                enc_inputs.append(make_input(f"ref_image_{si}", "IMAGE", link=None))
        else:
            enc_inputs.append(make_input(f"ref_image_{si}", "IMAGE", link=None))

    # CLIP 连接
    clip_lid = idgen.link()
    enc_inputs.insert(0, make_input("clip", "CLIP", link=clip_lid))
    links.append([clip_lid, shared_refs["clip_id"], shared_refs["clip_out"],
                  enc_id, 0, "CLIP"])

    # prompt 连接（STRING 是 widget 类型，连接时必须标注 widget 名称）
    prompt_lid = idgen.link()
    enc_inputs.insert(1, make_input("prompt", "STRING", link=prompt_lid, widget_name="prompt"))
    links.append([prompt_lid, pl_id, 0, enc_id, 1, "STRING"])

    enc_outputs = [
        make_output("conditioning", "CONDITIONING", links=[], slot=0),
    ]

    enc_node = make_node(
        enc_id, "H3EncodeConditioning", f"Encode ({shot_id})",
        [x_offset, y_offset], [400, 300],
        inputs=enc_inputs,
        outputs=enc_outputs,
        # widgets 顺序按 INPUT_TYPES 完整列表：prompt(被链接·占位), ref_image_short_edge
        widgets_values=["", 768],
    )
    nodes.append(enc_node)

    # ── H3SaveConditioning ──
    save_id = idgen.node()
    save_lid = idgen.link()
    links.append([save_lid, enc_id, 0, save_id, 0, "CONDITIONING"])

    # duration 连接 (per-shot PrimitiveFloat)
    save_dur_lid = idgen.link()
    links.append([save_dur_lid, dur_id, 0, save_id, 1, "FLOAT"])

    save_inputs = [
        make_input("conditioning", "CONDITIONING", link=save_lid),
        make_input("duration", "FLOAT", link=save_dur_lid, widget_name="duration"),
    ]

    # 参考图字节存入 .pt（供生成阶段 H3ReencodeFromCache 重新 VAE 编码）
    # 注意：每条连接必须使用唯一的 link ID，不能复用 H3EncodeConditioning 的 link
    # 且目标 slot 是 save_inputs 数组索引：conditioning=0, duration=1, ref_image_0~3=2~5
    save_ref_sources = {}
    for idx, (ln, _lid) in ref_image_links.items():
        if idx >= 4:
            break
        save_ref_sources[idx] = ln
    for idx in range(4):
        if idx in save_ref_sources:
            ln = save_ref_sources[idx]
            new_lid = idgen.link()
            save_inputs.append(make_input(f"ref_image_{idx}", "IMAGE", link=new_lid))
            links.append([new_lid, ln, 0, save_id, 2 + idx, "IMAGE"])
        else:
            save_inputs.append(make_input(f"ref_image_{idx}", "IMAGE", link=None))

    save_node = make_node(
        save_id, "H3SaveConditioning", f"Save ({shot_id}.pt)",
        [x_offset + 500, y_offset], [300, 180],
        inputs=save_inputs,
        outputs=[],
        # widgets 顺序按 INPUT_TYPES 完整列表（被链接的 widget 仍需占位！）：
        # filename, add_counter, duration(被链接·占位), width, height,
        # prompt, ref_image_size, ref_image_format
        # 前端按完整 widget 顺序应用 widgets_values 后才移除被链接 widget，
        # 缺占位会导致后续值整体错位（如 height 收到 ''、ref_image_size 收到 'jpeg'）
        widgets_values=[shot_id, True, 0, 0, 0, "", "match", "jpeg"],
        color="#232", bgcolor="#353",
    )
    nodes.append(save_node)

    return nodes, links


def generate_group_json(group_name, group_shots, asset_paths, output_dir):
    """为一个角色组生成多链 JSON。"""
    idgen = IDGen(start=1)
    all_nodes = []
    all_links = []

    # 从第一个镜头提取分辨率（全组共享）
    resolution = _extract_number(group_shots[0].get("分辨率", "0.5"), 0.5)

    # 共享节点
    shared_nodes, shared_refs = build_shared_nodes(idgen, resolution)
    all_nodes.extend(shared_nodes)

    # ── 合并 LoadImage ──
    # 收集本组所有用到的资产，去重
    load_nodes = {}  # (type, name) → node_id
    load_x = -1300
    load_y = 0

    # 主角色（全组共享）
    main_char = _clean_name(group_shots[0].get("参演角色1", ""))
    if main_char and main_char != "(纯场景)":
        path = asset_paths.get(("角色", main_char), f"{main_char}.png")
        nid, node = build_loadimage(idgen, main_char, path, load_x, load_y)
        all_nodes.append(node)
        load_nodes[("char", main_char)] = nid
        load_y += 100

    # 各场景（去重）——组内存在无场景镜头的，挂空 load 占位（__placeholder__），保证 ref 槽位存在
    placeholder_name, _created = _ensure_placeholder_image()
    seen_scenes = set()
    for shot in group_shots:
        for sk in ["场景设置1", "场景设置2"]:
            sname = _clean_name(shot.get(sk, ""))
            if sname and sname not in seen_scenes:
                seen_scenes.add(sname)
                path = asset_paths.get(("场景", sname), f"{sname}.png")
                nid, node = build_loadimage(idgen, sname, path, load_x, load_y)
                all_nodes.append(node)
                load_nodes[("scene", sname)] = nid
                load_y += 100
    has_scene_gap = any(not _clean_name(shot.get(sk, ""))
                        for shot in group_shots
                        for sk in ["场景设置1", "场景设置2"])
    if has_scene_gap and ("scene", "__placeholder__") not in load_nodes:
        # 有镜头未配置场景资产，挂占位 LoadImage 供用户后续替换
        nid, node = build_loadimage(idgen, "场景占位(未配置)", placeholder_name, load_x, load_y)
        all_nodes.append(node)
        load_nodes[("scene", "__placeholder__")] = nid
        load_y += 100

    # 各道具（去重）——组内存在无道具镜头的，挂空 load 占位（__placeholder__），保证 ref 槽位存在
    seen_props = set()
    for shot in group_shots:
        for pk in ["参演道具1", "参演道具2", "参演道具3"]:
            pname = _clean_name(shot.get(pk, ""))
            if pname and pname not in seen_props:
                seen_props.add(pname)
                path = asset_paths.get(("道具", pname), f"{pname}.png")
                nid, node = build_loadimage(idgen, pname, path, load_x, load_y)
                all_nodes.append(node)
                load_nodes[("prop", pname)] = nid
                load_y += 100
    has_prop_gap = any(not _clean_name(shot.get(pk, ""))
                       for shot in group_shots
                       for pk in ["参演道具1", "参演道具2", "参演道具3"])
    if has_prop_gap and ("prop", "__placeholder__") not in load_nodes:
        # 有镜头未配置道具资产，挂占位 LoadImage 供用户后续替换
        nid, node = build_loadimage(idgen, "道具占位(未配置)", placeholder_name, load_x, load_y)
        all_nodes.append(node)
        load_nodes[("prop", "__placeholder__")] = nid
        load_y += 100

    # 配角（去重）
    seen_sups = set()
    for shot in group_shots:
        for ck in ["参演角色2", "参演角色3"]:
            cname = _clean_name(shot.get(ck, ""))
            if cname and cname != main_char and cname not in seen_sups:
                seen_sups.add(cname)
                path = asset_paths.get(("角色", cname), f"{cname}.png")
                nid, node = build_loadimage(idgen, cname, path, load_x, load_y)
                all_nodes.append(node)
                load_nodes[("char", cname)] = nid
                load_y += 100

    # 参考音频（去重）
    seen_audios = set()
    for shot in group_shots:
        for ak in ["参考音频1", "参考音频2"]:
            aname = _clean_name(shot.get(ak, ""))
            if aname and aname not in seen_audios:
                seen_audios.add(aname)
                path = asset_paths.get(("音频", aname), "")
                nid, node = build_loadaudio(idgen, aname, path, load_x, load_y)
                all_nodes.append(node)
                load_nodes[("audio", aname)] = nid
                load_y += 100

    # 参考视频（去重）
    seen_videos = set()
    for shot in group_shots:
        vname = _clean_name(shot.get("参考视频", ""))
        if vname and vname not in seen_videos:
            seen_videos.add(vname)
            path = asset_paths.get(("视频", vname), "")
            nid, node = build_loadimage(idgen, vname, path, load_x, load_y)
            all_nodes.append(node)
            load_nodes[("video", vname)] = nid
            load_y += 100

    # ── 逐镜创建链 ──
    for i, shot in enumerate(group_shots):
        x_off = -500 + (i % 4) * 700
        y_off = -420 + (i // 4) * 500
        chain_nodes, chain_links = build_shot_chain(
            idgen, shot, shared_refs, load_nodes, asset_paths, x_off, y_off
        )
        all_nodes.extend(chain_nodes)
        all_links.extend(chain_links)

    # 更新共享节点的输出 links 列表
    # CLIPLoader 输出
    for node in all_nodes:
        if node["type"] == "CLIPLoader" and node["outputs"]:
            clip_links = [l[0] for l in all_links if l[1] == node["id"] and l[2] == 0]
            node["outputs"][0]["links"] = clip_links
        elif node["type"] == "H3EncodeConditioning" and node["outputs"]:
            enc_links = [l[0] for l in all_links if l[1] == node["id"] and l[2] == 0]
            node["outputs"][0]["links"] = enc_links
        elif node["type"] == "LoadImage":
            img_links = [l[0] for l in all_links if l[1] == node["id"] and l[2] == 0]
            node["outputs"][0]["links"] = img_links
        elif node["type"] == "LoadAudio":
            aud_links = [l[0] for l in all_links if l[1] == node["id"] and l[2] == 0]
            node["outputs"][0]["links"] = aud_links
        elif node["type"] == "PrimitiveFloat":
            fl_links = [l[0] for l in all_links if l[1] == node["id"] and l[2] == 0]
            node["outputs"][0]["links"] = fl_links
        elif node["type"] in ("PrimitiveStringMultiline", "PrimitiveString"):
            pl_links = [l[0] for l in all_links if l[1] == node["id"] and l[2] == 0]
            if node["outputs"]:
                node["outputs"][0]["links"] = pl_links

    # 构建 JSON
    wf = {
        "id": f"denghuang-h3-preencode-{group_name}",
        "revision": 0,
        "last_node_id": idgen._next - 1,
        "last_link_id": idgen._next - 1,
        "nodes": all_nodes,
        "links": all_links,
        "groups": [
            {
                "id": 1,
                "title": f"{group_name} ({len(group_shots)}镜)",
                "bounding": [-1850, -480, 2200, 1200],
                "color": "#3f789e",
                "flags": {},
            }
        ],
        "config": {},
        "extra": {
            "ds": {"scale": 0.6, "offset": [800, 400]},
            "note": f"H3 多链预编码 | {group_name} | {len(group_shots)}镜 | 共享Load合并 | "
                    f"每镜独立分辨率+时长 | 元数据连接",
        },
        "version": 0.4,
    }

    safe_name = sanitize_filename(group_name)
    output_path = os.path.join(output_dir, f"{safe_name}.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(wf, f, ensure_ascii=False, indent=1)

    return output_path, len(group_shots), len(load_nodes)


# ── 资产映射 ─────────────────────────────────────────────────────────

def load_asset_mapping(mapping_path):
    """从 JSON 文件加载资产映射，返回 {(type, name): path} dict。
    JSON 格式: {"角色": {"黑猫": "path.png", ...}, "场景": {...}, "道具": {...}}
    """
    if not os.path.exists(mapping_path):
        return {}
    with open(mapping_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    mapping = {}
    for asset_type, items in data.items():
        if isinstance(items, dict):
            for name, path in items.items():
                if path:
                    mapping[(asset_type, name)] = path
    return mapping


def merge_asset_paths(excel_paths, mapping_paths):
    """合并 Excel Sheet2 路径和外部映射路径。
    Excel Sheet2 优先，映射补充空缺。
    """
    merged = dict(mapping_paths)
    merged.update(excel_paths)
    return merged


# ── 单文件 / 批量处理 ───────────────────────────────────────────────

def process_single(xlsx_path, output_dir, by_shot=False, log=print,
                   asset_mapping=None):
    """处理单个 Excel 文件，返回 (n_groups, n_shots) 或 None。
    asset_mapping: 可选，{(type, name): path} dict，补充 Excel Sheet2 的空缺。
    """
    if not Path(xlsx_path).exists():
        log(f"ERROR: file not found: {xlsx_path}")
        return None

    os.makedirs(output_dir, exist_ok=True)
    shots, excel_paths = read_excel(xlsx_path)

    if not shots:
        log(f"ERROR: no shots in {xlsx_path}")
        return None

    # 合并资产路径：Excel Sheet2 优先，映射补充
    if asset_mapping:
        asset_paths = merge_asset_paths(excel_paths, asset_mapping)
        log(f"  读取: {len(shots)} 镜, Excel {len(excel_paths)} 路径 + 映射 {len(asset_mapping)} 路径")
    else:
        asset_paths = excel_paths
        log(f"  读取: {len(shots)} 镜, {len(asset_paths)} 个资产路径")

    # 分组
    if by_shot:
        groups = {}
        for shot in shots:
            sid = shot.get("镜头编号", "unknown")
            groups[sid] = [shot]
    else:
        groups = {}
        for shot in shots:
            char = _clean_name(shot.get("参演角色1", "(纯场景)"))
            if not char:
                char = "(纯场景)"
            groups.setdefault(char, []).append(shot)

    log(f"  分组: {len(groups)} 组")
    for name, group in groups.items():
        path, n_shots, n_loads = generate_group_json(name, group, asset_paths, output_dir)
        log(f"    {name}: {n_shots}镜, {n_loads}个Load -> {os.path.basename(path)}")

    return len(groups), len(shots)


def process_batch(xlsx_files, output_dir, by_shot=False, log=print,
                  asset_mapping=None):
    """批量处理多个 Excel 文件。每个 Excel 生成到 output_dir 下的子目录。"""
    success, fail = 0, 0

    for xlsx_path in xlsx_files:
        stem = Path(xlsx_path).stem
        sub_dir = os.path.join(output_dir, stem)
        log(f"\n[{success + fail + 1}/{len(xlsx_files)}] {Path(xlsx_path).name}")

        result = process_single(xlsx_path, sub_dir, by_shot=by_shot, log=log,
                                asset_mapping=asset_mapping)
        if result:
            success += 1
        else:
            fail += 1

    log(f"\n批量完成: {success} 成功, {fail} 失败, 共 {len(xlsx_files)} 文件")
    return success, fail


# ── 主入口 ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="提示词审阅表 → 多链生产 JSON")
    parser.add_argument("files", nargs="+", help="Excel 文件路径（一个或多个）")
    parser.add_argument("-o", "--output", required=True, help="输出目录")
    parser.add_argument("--by-shot", action="store_true",
                        help="按镜头顺序分组（每镜一个JSON）")
    parser.add_argument("-m", "--mapping", help="资产映射 JSON 文件路径（补充 Excel Sheet2 空缺）")
    args = parser.parse_args()

    xlsx_files = args.files
    asset_mapping = None
    if args.mapping:
        asset_mapping = load_asset_mapping(args.mapping)
        if asset_mapping:
            print(f"加载资产映射: {len(asset_mapping)} 条 from {args.mapping}")
        else:
            print(f"WARNING: 资产映射为空或文件不存在: {args.mapping}")

    if len(xlsx_files) == 1:
        result = process_single(xlsx_files[0], args.output, by_shot=args.by_shot,
                                asset_mapping=asset_mapping)
        if result is None:
            sys.exit(1)
    else:
        process_batch(xlsx_files, args.output, by_shot=args.by_shot,
                      asset_mapping=asset_mapping)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
shot_md_to_excel.py
===================
MD → 提示词审阅表 Excel

支持三种格式：
  - 分镜头需求（旧格式 ### A01 + H3提示词）
  - 分镜头需求 v6（#### 镜头N + 10字段，不含 H3 提示词）
  - H3 提示词（## A01 九分节格式）

功能：
  - 格式自动检测 + 解析
  - 提示词审阅表 / 分镜审阅表 Excel 生成
  - 规范自检（指代不明 / 对白格式 / 画风冲突 / 时长一致性）

用法:
  单文件:  python shot_md_to_excel.py <input.md> <output.xlsx>
  批量:    python shot_md_to_excel.py -o <输出目录> <file1.md> [file2.md ...]
"""

import re
import sys
import os
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


# ── MD 解析 ──────────────────────────────────────────────────────────

def _split_kv_line(line):
    """拆分一行中可能含多个 |**key**: val 的键值对。返回 [(key, val), ...]"""
    m = re.match(r"-\s*\*\*(.+?)\*\*:\s*(.*)$", line)
    if not m:
        return []
    key, rest = m.group(1), m.group(2)
    results = []
    while True:
        m2 = re.match(r"(.*?)\s*\|\s*\*\*([^*]+)\*\*:\s*(.*)$", rest, re.DOTALL)
        if not m2:
            results.append((key, rest.strip()))
            break
        cur_val, nk, nrest = m2.group(1), m2.group(2), m2.group(3)
        results.append((key, cur_val.strip()))
        key, rest = nk, nrest
    return results


def _parse_global_info(text):
    """从 ## 全局信息 区块解析键值对（兼容 `- **键**: 值` 与 `- 键: 值`）。"""
    global_info = {}
    g_match = re.search(r"## 全局信息\s*\n(.*?)(?=## 镜头列表)", text, re.DOTALL)
    if g_match:
        for line in g_match.group(1).strip().splitlines():
            m = re.match(r"-\s*\*\*(.+?)\*\*:\s*(.+)", line)
            if not m:
                m = re.match(r"-\s*([^:：]+?)[:：]\s*(.+)", line)
            if m:
                global_info[m.group(1).strip()] = m.group(2).strip()
    return global_info


def parse_md(md_path):
    """解析分镜头需求 MD 文件（旧格式 ### A01 + H3提示词），返回 (global_info, shots)。"""
    text = Path(md_path).read_text(encoding="utf-8")

    global_info = _parse_global_info(text)

    shots = []
    parts = re.split(r"^### ", text, flags=re.MULTILINE)
    for part in parts:
        part = part.strip()
        if not part or part.startswith("#"):
            continue
        lines = part.splitlines()
        shot_id = lines[0].strip()
        if not re.match(r"^[A-Z]\d+", shot_id):
            continue

        shot = {"id": shot_id}
        for line in lines[1:]:
            line = line.strip()
            for key, val in _split_kv_line(line):
                shot[key] = val

        prompt_match = re.search(
            r"#### 完整 H3 提示词\s*\n(.*?)(?=#### |^---|\Z)", part, re.DOTALL
        )
        if prompt_match:
            shot["完整提示词"] = prompt_match.group(1).strip()

        shots.append(shot)

    return global_info, shots


def parse_storyboard_v6(md_path):
    """解析 v6 结构化分镜表（#### 镜头N + 10字段，不含 H3 提示词）。

    返回 (global_info, shots)。
    """
    text = Path(md_path).read_text(encoding="utf-8", errors="replace")

    global_info = _parse_global_info(text)

    shots = []
    parts = re.split(r"^#### ", text, flags=re.MULTILINE)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        lines = part.splitlines()
        shot_id = lines[0].strip()
        if not re.match(r"^镜头\d+", shot_id):
            continue

        shot = {"id": shot_id}
        for line in lines[1:]:
            line = line.strip()
            for key, val in _split_kv_line(line):
                shot[key] = val
        shots.append(shot)

    return global_info, shots


# ── H3 提示词 MD 解析（## A01 九分节格式）─────────────────────────────

H3_SECTIONS = [
    "输出规格", "参考图约束", "整体风格", "场景描述", "两级时间轴",
    "摄影与摄像机", "光影", "声音", "约束条件",
]


def detect_format(md_path):
    """检测 MD 文件格式。

    返回 "storyboard"（旧格式 ### A01 + H3提示词）或
          "storyboard_v6"（新格式 #### 镜头N + 10字段，不含 H3 提示词）或
          "h3_prompt"（H3提示词，## A01 九分节）或 None。
    """
    text = Path(md_path).read_text(encoding="utf-8", errors="replace")
    if re.search(r"^##\s+(?:[A-Z]\d+|镜头\d+)\s*[-—]", text, re.MULTILINE):
        return "h3_prompt"
    if re.search(r"^####\s+镜头\d+", text, re.MULTILINE):
        return "storyboard_v6"
    if re.search(r"^###\s+[A-Z]\d+", text, re.MULTILINE):
        return "storyboard"
    return None


def parse_h3_prompt_md(md_path):
    """解析 H3 提示词 MD 文件（## A01 九分节格式）。

    返回 (meta, shots)。meta 为文件头信息 dict；shots 为镜头 dict 列表，
    每镜含 id / 名称 / 时长 / 各分节内容 / 完整提示词。
    """
    text = Path(md_path).read_text(encoding="utf-8", errors="replace")

    # 文件头元信息（第一个 ## 之前）
    meta = {}
    head = text.split("\n## ", 1)[0]
    for line in head.splitlines():
        m = re.match(r">\s*(.+?)[：:]\s*(.+)", line.strip())
        if m:
            meta[m.group(1).strip()] = m.group(2).strip()

    shots = []
    parts = re.split(r"^## ", text, flags=re.MULTILINE)
    for part in parts[1:]:
        part = part.strip()
        if not part:
            continue
        lines = part.splitlines()
        title = lines[0].strip()
        # 兼容 ## A01 与 ## 镜头N（镜头N 映射为 A编号，如 镜头1→A01、镜头10→A10、镜头243→A243）
        m = re.match(r"^([A-Z]\d+)\s*[-—]\s*(.*)$", title)
        if not m:
            m = re.match(r"^镜头(\d+)\s*[-—]\s*(.*)$", title)
            if m:
                shot_id, shot_name = f"A{int(m.group(1)):02d}", m.group(2).strip()
            else:
                continue
        else:
            shot_id, shot_name = m.group(1), m.group(2).strip()

        shot = {"id": shot_id, "名称": shot_name}
        body = "\n".join(lines[1:])

        # 按九分节切分
        section_pat = re.compile(r"^【(.+?)】\s*$", re.MULTILINE)
        matches = list(section_pat.finditer(body))
        for i, m in enumerate(matches):
            sec_name = m.group(1)
            end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
            content = body[m.end():end].strip()
            shot[sec_name] = content

        # 从【输出规格】提取时长
        spec = shot.get("输出规格", "")
        dur_m = re.search(r"(\d+(?:\.\d+)?)\s*秒", spec)
        if dur_m:
            shot["时长"] = dur_m.group(1)

        shot["完整提示词"] = body
        shots.append(shot)

    return meta, shots


def _h3_asset_from_refs(refs_text, kind):
    """从【参考图约束】中提取指定类型（场景/角色/道具）的资产名。

    兼容两种格式：
    - 类型词前置：`- <图片1> = 场景参考图：金山町山谷...`（场景/道具）
    - 角色名前置：`- <图片2> = 金止戈角色参考图：...` / `- <图片2> = 金止戈（剪影）角色参考图：...`（角色）
    kind 为 "场景"/"角色"/"道具"。
    """
    if not refs_text:
        return []
    names = []
    is_role = (kind == "角色")
    for line in refs_text.splitlines():
        line = line.strip()
        if not line.startswith("-"):
            continue
        if is_role:
            # 角色名在"角色参考图"前：`= X角色参考图：`
            m = re.search(r"=\s*([^：:]+?)\s*角色参考图[：:]", line)
            if m:
                name = _strip_bracket(m.group(1).strip())
                if name and name not in names:
                    names.append(name)
        else:
            # 类型词前置：`= 场景参考图：名称 ...` / `= 道具参考图：名称 ...`
            m = re.search(rf"=\s*{kind}参考图[：:]\s*([^，,、。．；;]+)", line)
            if m:
                name = _strip_bracket(m.group(1).strip())
                if name and name not in names:
                    names.append(name)
    return names


def _strip_bracket(s):
    """去掉名称中的括号注释：`金止戈（剪影）` → `金止戈`；`M48坦克（左1/3）` → `M48坦克`。
    兼容未闭合括号：`金止戈大刀（说明` → `金止戈大刀`；`布口袋）` → `布口袋`。"""
    s = re.sub(r"[（(][^（）()]*[）)]", "", s)   # 成对括号含内容
    s = re.sub(r"[（(][^）()]*$", "", s)          # 行尾未闭合左括号段
    s = re.sub(r"[）)]+$", "", s)                 # 行尾孤立右括号
    return s.strip()


def _h3_parse_time_axis(text):
    """从【两级时间轴】提取 (总时长, 时间轴文本)。"""
    if not text:
        return "", ""
    first = text.splitlines()[0]
    m = re.search(r"时长\s*([\d.]+)\s*秒", first)
    duration = m.group(1) if m else ""
    return duration, text


# ── Excel 生成 ───────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="3F789E", end_color="3F789E", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# 对白红色字体（用于 <d></d> 标记的台词部分）
DIALOGUE_FONT = InlineFont(color="FF0000")
NORMAL_FONT = InlineFont(color="000000")

SHEET1_HEADERS = [
    "镜头编号", "时长", "分辨率", "场景", "镜头调度", "约束条件", "镜头风格",
    "完整提示词",
    "参演角色1", "参演角色2", "参演角色3",
    "场景设置1", "场景设置2",
    "参演道具1", "参演道具2", "参演道具3",
    "参考音频1", "参考音频2", "参考视频",
    "修改指令",
]

SHEET2_HEADERS = ["类型", "名称", "input路径"]


def _split_assets(val):
    if not val or val == "(无)":
        return []
    # 同时按逗号和顿号拆分（道具常含顿号），并清理 "| **键**: 值" 残留
    raw_parts = re.split(r"[,、]", val)
    out = []
    for s in raw_parts:
        s = s.strip()
        if not s or s in ("(无)", "(黑场)"):
            continue
        # 去掉 "| **场景参考**: xxx" 这类残留
        s = re.split(r"\|\s*\*\*", s)[0].strip()
        # 清理角色括号前缀如 "(剪影·金止戈)" -> "金止戈"、"(群演·游击队员)" -> "游击队员"
        # 兼容多角色同括号被顿号切开的情况（如 "(剪影·金止戈、" 与 "赤丸冢羽)"）
        s = re.sub(r"^\((?:剪影|群演)?[··]?\s*", "", s)
        s = s.rstrip(")")
        if s:
            out.append(s)
    return out


def _col_letter(col_idx):
    if col_idx <= 26:
        return chr(64 + col_idx)
    return chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)


def _make_rich_text(text):
    """将包含 <d></d> 标记的文本转为 CellRichText，台词部分红色显示。

    无 <d> 标记时返回原始字符串（不创建 CellRichText，避免开销）。
    """
    if not text or "<d>" not in str(text):
        return text

    text = str(text)
    blocks = []
    pattern = re.compile(r"(<d>.*?</d>)", re.DOTALL)
    last_end = 0

    for m in pattern.finditer(text):
        # 标记前的普通文本
        if m.start() > last_end:
            blocks.append(TextBlock(NORMAL_FONT, text[last_end:m.start()]))
        # <d></d> 内的台词文本（红色）
        blocks.append(TextBlock(DIALOGUE_FONT, m.group(1)))
        last_end = m.end()

    # 末尾普通文本
    if last_end < len(text):
        blocks.append(TextBlock(NORMAL_FONT, text[last_end:]))

    if not blocks:
        return text

    return CellRichText(blocks)


def generate_excel(global_info, shots, output_path):
    """生成 Excel 审阅表。返回 (n_shots, n_assets) 元组。"""
    wb = Workbook()

    # ── Sheet1: 提示词审阅 ──
    ws1 = wb.active
    ws1.title = "提示词审阅"

    for col, header in enumerate(SHEET1_HEADERS, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, shot in enumerate(shots, 2):
        characters = _split_assets(shot.get("参演角色", ""))
        scene_refs = _split_assets(shot.get("场景参考", ""))
        props = _split_assets(shot.get("道具参考", ""))

        if not characters and scene_refs:
            characters = ["(纯场景)"]

        row_data = [
            shot["id"],
            shot.get("时长", ""),
            shot.get("分辨率", "0.5"),
            shot.get("场景 / Scene", shot.get("场景", "")),
            shot.get("镜头调度 / Camera", shot.get("镜头调度", "")),
            shot.get("约束条件 / Constraints", shot.get("约束条件", "")),
            shot.get("镜头风格 / Style", shot.get("镜头风格", "")),
            shot.get("完整提示词", ""),
            characters[0] if len(characters) > 0 else "",
            characters[1] if len(characters) > 1 else "",
            characters[2] if len(characters) > 2 else "",
            scene_refs[0] if len(scene_refs) > 0 else "",
            scene_refs[1] if len(scene_refs) > 1 else "",
            props[0] if len(props) > 0 else "",
            props[1] if len(props) > 1 else "",
            props[2] if len(props) > 2 else "",
            shot.get("参考音频1", ""),
            shot.get("参考音频2", ""),
            shot.get("参考视频", ""),
            "",
        ]

        for col, value in enumerate(row_data, 1):
            # 第8列"完整提示词"：对 <d></d> 标记的台词部分用红色富文本显示
            if col == 8:
                value = _make_rich_text(value)
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = {
        1: 10, 2: 8, 3: 10, 4: 30, 5: 30, 6: 25, 7: 20,
        8: 60, 9: 12, 10: 12, 11: 12,
        12: 15, 13: 15, 14: 12, 15: 12, 16: 12,
        17: 15, 18: 15, 19: 15, 20: 30,
    }
    for col, width in col_widths.items():
        ws1.column_dimensions[_col_letter(col)].width = width
    ws1.freeze_panes = "A2"

    # 分辨率列下拉验证 (C列, 0.4 / 0.5)
    res_dv = DataValidation(type="list", formula1='"0.4,0.5"', allow_blank=False)
    res_dv.error = "请选择 0.4 或 0.5"
    res_dv.errorTitle = "无效分辨率"
    res_dv.prompt = "H3分辨率系数: 0.5=高清, 0.4=经济"
    res_dv.promptTitle = "分辨率选择"
    ws1.add_data_validation(res_dv)
    for row_idx in range(2, len(shots) + 2):
        res_dv.add(ws1.cell(row=row_idx, column=3))

    # ── Sheet2: 美术资产路径 ──
    ws2 = wb.create_sheet("美术资产路径")
    for col, header in enumerate(SHEET2_HEADERS, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    asset_set = set()
    for shot in shots:
        for c in _split_assets(shot.get("参演角色", "")):
            asset_set.add(("角色", c))
        for s in _split_assets(shot.get("场景参考", "")):
            asset_set.add(("场景", s))
        for p in _split_assets(shot.get("道具参考", "")):
            asset_set.add(("道具", p))
        for a in _split_assets(shot.get("参考音频1", "")):
            asset_set.add(("音频", a))
        for a in _split_assets(shot.get("参考音频2", "")):
            asset_set.add(("音频", a))
        for v in _split_assets(shot.get("参考视频", "")):
            asset_set.add(("视频", v))

    type_order = {"角色": 0, "场景": 1, "道具": 2, "音频": 3, "视频": 4}
    sorted_assets = sorted(asset_set, key=lambda x: (type_order.get(x[0], 9), x[1]))

    for row_idx, (asset_type, asset_name) in enumerate(sorted_assets, 2):
        ws2.cell(row=row_idx, column=1, value=asset_type).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=asset_name).border = THIN_BORDER
        ws2.cell(row=row_idx, column=3, value="").border = THIN_BORDER

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 70
    ws2.freeze_panes = "A2"

    # ── Sheet3: 说明 ──
    ws3 = wb.create_sheet("说明")
    notes = [
        "提示词审阅表使用说明",
        "",
        "1. Sheet1「提示词审阅」",
        "   - 每行一镜，从分镜头需求 MD 自动提取",
        "   - 时长 列可审阅修改（单位: 秒，直接编辑数值）",
        "   - 分辨率 列下拉选择 0.4 或 0.5（对应 H3 节点分辨率系数，0.5=高清, 0.4=经济）",
        "   - 场景/镜头调度/约束条件/镜头风格 列从 H3 提示词要素提取",
        "   - 完整提示词 列为 H3 三核心字段原文，其中 <d>[语言] 台词</d> 标记的台词部分以红色显示",
        "   - 参演角色1-3 / 场景设置1-2 / 参演道具1-3 从美术资产需求提取",
        "   - 参考音频1-2 / 参考视频 从 H3 提示词要素提取（可选，留空则不生成对应 Load 节点）",
        "   - 修改指令 列留空，供审阅时填写修改意见",
        "",
        "2. Sheet2「美术资产路径」",
        "   - 自动收集所有镜头涉及的资产（角色/场景/道具/音频/视频），去重排列",
        "   - 在 input路径 列填写对应的 ComfyUI input 目录相对路径",
        "   - 示例: h3_ref/角色/黑猫沈天然/沈天然黑猫定妆照.png",
        "",
        "3. 审阅完成后",
        "   - 保存 Excel，运行 excel_to_multi_chain_json.py 生成多链生产 JSON",
        "   - 用法: python excel_to_multi_chain_json.py <审阅表.xlsx> <输出目录>",
    ]
    for row_idx, note in enumerate(notes, 1):
        ws3.cell(row=row_idx, column=1, value=note)
    ws3.column_dimensions["A"].width = 80

    wb.save(output_path)
    return len(shots), len(sorted_assets)


H3_SHEET1_HEADERS = [
    "镜头编号", "镜头名称", "时长", "分辨率", "场景", "镜头调度", "约束条件", "镜头风格",
    "完整提示词",
    "参演角色1", "参演角色2", "参演角色3",
    "场景设置1", "场景设置2",
    "参演道具1", "参演道具2", "参演道具3",
    "参考音频1", "参考音频2", "参考视频",
    "修改指令",
]


def generate_h3_prompt_excel(meta, shots, output_path):
    """生成 H3 提示词审阅表 Excel。返回 (n_shots, n_assets)。"""
    wb = Workbook()

    # ── Sheet1: 提示词审阅 ──
    ws1 = wb.active
    ws1.title = "提示词审阅"

    for col, header in enumerate(H3_SHEET1_HEADERS, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, shot in enumerate(shots, 2):
        refs = shot.get("参考图约束", "")
        characters = _h3_asset_from_refs(refs, "角色")
        scene_refs = _h3_asset_from_refs(refs, "场景")
        props = _h3_asset_from_refs(refs, "道具")

        if not characters and scene_refs:
            characters = ["(纯场景)"]

        axis_dur, _ = _h3_parse_time_axis(shot.get("两级时间轴", ""))
        duration = shot.get("时长") or axis_dur

        row_data = [
            shot["id"],
            shot.get("名称", ""),
            duration,
            "0.5",
            shot.get("场景描述", ""),
            shot.get("摄影与摄像机", ""),
            shot.get("约束条件", ""),
            shot.get("整体风格", ""),
            shot.get("完整提示词", ""),
            characters[0] if len(characters) > 0 else "",
            characters[1] if len(characters) > 1 else "",
            characters[2] if len(characters) > 2 else "",
            scene_refs[0] if len(scene_refs) > 0 else "",
            scene_refs[1] if len(scene_refs) > 1 else "",
            props[0] if len(props) > 0 else "",
            props[1] if len(props) > 1 else "",
            props[2] if len(props) > 2 else "",
            "", "", "",
            "",
        ]

        for col, value in enumerate(row_data, 1):
            if col == 9:
                value = _make_rich_text(value)
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = {
        1: 10, 2: 20, 3: 8, 4: 10, 5: 40, 6: 40, 7: 35, 8: 25,
        9: 60, 10: 12, 11: 12, 12: 12,
        13: 15, 14: 15, 15: 12, 16: 12, 17: 12,
        18: 15, 19: 15, 20: 15, 21: 30,
    }
    for col, width in col_widths.items():
        ws1.column_dimensions[_col_letter(col)].width = width
    ws1.freeze_panes = "A2"

    res_dv = DataValidation(type="list", formula1='"0.4,0.5"', allow_blank=False)
    res_dv.error = "请选择 0.4 或 0.5"
    res_dv.errorTitle = "无效分辨率"
    res_dv.prompt = "H3分辨率系数: 0.5=高清, 0.4=经济"
    res_dv.promptTitle = "分辨率选择"
    ws1.add_data_validation(res_dv)
    for row_idx in range(2, len(shots) + 2):
        res_dv.add(ws1.cell(row=row_idx, column=4))

    # ── Sheet2: 美术资产路径 ──
    ws2 = wb.create_sheet("美术资产路径")
    for col, header in enumerate(SHEET2_HEADERS, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    asset_set = set()
    for shot in shots:
        refs = shot.get("参考图约束", "")
        for c in _h3_asset_from_refs(refs, "角色"):
            asset_set.add(("角色", c))
        for s in _h3_asset_from_refs(refs, "场景"):
            asset_set.add(("场景", s))
        for p in _h3_asset_from_refs(refs, "道具"):
            asset_set.add(("道具", p))

    type_order = {"角色": 0, "场景": 1, "道具": 2, "音频": 3, "视频": 4}
    sorted_assets = sorted(asset_set, key=lambda x: (type_order.get(x[0], 9), x[1]))

    for row_idx, (asset_type, asset_name) in enumerate(sorted_assets, 2):
        ws2.cell(row=row_idx, column=1, value=asset_type).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=asset_name).border = THIN_BORDER
        ws2.cell(row=row_idx, column=3, value="").border = THIN_BORDER

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 70
    ws2.freeze_panes = "A2"

    # ── Sheet3: 说明 ──
    ws3 = wb.create_sheet("说明")
    notes = [
        "H3 提示词审阅表使用说明",
        "",
        "1. Sheet1「提示词审阅」",
        "   - 每行一镜，从 H3 提示词 MD（## A01 九分节格式）自动提取",
        "   - 时长 列可审阅修改（单位: 秒，直接编辑数值）",
        "   - 分辨率 列下拉选择 0.4 或 0.5（对应 H3 节点分辨率系数，0.5=高清, 0.4=经济）",
        "   - 场景/镜头调度/约束条件/镜头风格 列分别取自【场景描述】【摄影与摄像机】【约束条件】【整体风格】",
        "   - 完整提示词 列为九分节原文，其中 <d>[语言][性别] 台词</d> 标记的台词部分以红色显示",
        "   - 参演角色/场景设置/参演道具 从【参考图约束】按类型提取",
        "   - 修改指令 列留空，供审阅时填写修改意见",
        "",
        "2. Sheet2「美术资产路径」",
        "   - 自动收集所有镜头涉及的资产（角色/场景/道具），去重排列",
        "   - 在 input路径 列填写对应的 ComfyUI input 目录相对路径",
        "   - 示例: h3_ref/角色/金止戈/金止戈定妆照.png",
        "",
        "3. 审阅完成后",
        "   - 保存 Excel，运行 excel_to_multi_chain_json.py 生成多链生产 JSON",
        "   - 用法: python excel_to_multi_chain_json.py <审阅表.xlsx> <输出目录>",
    ]
    for row_idx, note in enumerate(notes, 1):
        ws3.cell(row=row_idx, column=1, value=note)
    ws3.column_dimensions["A"].width = 80

    wb.save(output_path)
    return len(shots), len(sorted_assets)


V6_SHEET1_HEADERS = [
    "镜头编号", "镜头名称", "时长", "景别", "机位角度", "运镜", "构图",
    "画面内容", "人物动作", "声音", "转场", "Hook",
    "承接", "交接", "空间锚点卡", "对白时长核算", "修改指令",
]


def _v6_duration(val):
    m = re.search(r"(\d+(?:\.\d+)?)\s*秒", val or "")
    return m.group(1) if m else ""


def generate_storyboard_v6_excel(global_info, shots, output_path):
    """生成 v6 分镜审阅表 Excel。返回 (n_shots, n_assets)。"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "分镜审阅"

    for col, header in enumerate(V6_SHEET1_HEADERS, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, shot in enumerate(shots, 2):
        row_data = [
            shot["id"],
            shot.get("镜头名称", ""),
            _v6_duration(shot.get("时长", "")),
            shot.get("景别", ""),
            shot.get("机位角度", ""),
            shot.get("运镜", ""),
            shot.get("构图", ""),
            shot.get("画面内容", ""),
            shot.get("人物动作", ""),
            shot.get("声音", ""),
            shot.get("转场", ""),
            shot.get("Hook", ""),
            shot.get("承接 S(N-1)", ""),
            shot.get("交接 S(N+1)", ""),
            shot.get("空间锚点卡", ""),
            shot.get("对白时长核算", ""),
            "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = {
        1: 10, 2: 22, 3: 8, 4: 10, 5: 12, 6: 14, 7: 20,
        8: 40, 9: 40, 10: 30, 11: 12, 12: 10,
        13: 22, 14: 22, 15: 45, 16: 22, 17: 30,
    }
    for col, width in col_widths.items():
        ws1.column_dimensions[_col_letter(col)].width = width
    ws1.freeze_panes = "A2"

    # ── Sheet2: 说明 ──
    ws2 = wb.create_sheet("说明")
    notes = [
        "v6 分镜审阅表使用说明",
        "",
        "1. Sheet1「分镜审阅」",
        "   - 每行一镜，从 v6 结构化分镜表（#### 镜头N + 10字段）自动提取",
        "   - 本表为分镜头需求审阅，不含 H3 提示词（已拆分至 H3提示词 MD）",
        "   - 时长 列可审阅修改（单位: 秒，直接编辑数值）",
        "   - 修改指令 列留空，供审阅时填写修改意见",
        "",
        "2. H3 提示词审阅",
        "   - 请将 H3提示词 MD 文件（## A01 九分节）加入本工具生成提示词审阅表",
    ]
    for row_idx, note in enumerate(notes, 1):
        ws2.cell(row=row_idx, column=1, value=note)
    ws2.column_dimensions["A"].width = 80

    wb.save(output_path)
    return len(shots), 0


def collect_assets_from_md(md_path):
    """从单个 MD 文件收集资产清单，返回 dict: {type: set(names)}。
    type 为 "角色"/"场景"/"道具"/"音频"/"视频"。
    """
    _, shots = parse_md(md_path)
    assets = {"角色": set(), "场景": set(), "道具": set(), "音频": set(), "视频": set()}
    for shot in shots:
        for c in _split_assets(shot.get("参演角色", "")):
            assets["角色"].add(c)
        for s in _split_assets(shot.get("场景参考", "")):
            assets["场景"].add(s)
        for p in _split_assets(shot.get("道具参考", "")):
            assets["道具"].add(p)
        for a in _split_assets(shot.get("参考音频1", "")):
            assets["音频"].add(a)
        for a in _split_assets(shot.get("参考音频2", "")):
            assets["音频"].add(a)
        for v in _split_assets(shot.get("参考视频", "")):
            assets["视频"].add(v)
    return assets


def collect_assets_batch(md_files):
    """从多个 MD 文件收集全局资产清单。
    返回 dict: {"global": {type: set}, "episodes": {stem: {type: set}}}
    """
    global_assets = {"角色": set(), "场景": set(), "道具": set(), "音频": set(), "视频": set()}
    episodes = {}
    for md_path in md_files:
        stem = Path(md_path).stem
        ep_assets = collect_assets_from_md(md_path)
        episodes[stem] = ep_assets
        for t in global_assets:
            global_assets[t] |= ep_assets[t]
    return {"global": global_assets, "episodes": episodes}


def save_asset_mapping(mapping, output_path):
    """将资产映射保存为 JSON 文件。
    mapping 格式: {"角色": {"黑猫": "path/to.png", ...}, "场景": {...}, "道具": {...}}
    """
    import json
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


# ── 规范自检 ──────────────────────────────────────────────────────────

# 指代不明黑名单（无主指代，含剪影/身影等笼统指代）
VAGUE_REF_PATTERNS = [
    r"人影", r"一个人", r"那个人", r"这个人", r"某个人", r"一个人影",
    r"身影", r"一个身影", r"模糊的人", r"剪影中的人", r"人影幢幢",
]

# 画风锚点关键词（【整体风格】需含其一）
STYLE_ANCHOR_KEYS = ["D4rkL1nes", "画风锚点", "锚点", "风格锚点"]

# 画风冲突关键词（2D赛璐璐风格下不应出现）
STYLE_CONFLICT_KEYS = [
    "实拍", "3D渲染", "写实照片", "照片级", "photorealistic",
    "3D质感", "CG渲染", "实景", "实拍质感",
]

# 非叙事音乐规范句关键词（【声音】非叙事音乐需含其一）
MUSIC_RULE_KEYS = ["无背景音乐", "禁止任何配乐", "不要背景音乐", "仅保留环境音"]

# 对白语言标签
LANG_TAGS = ["Chinese", "Japanese", "English", "Korean", "Mixed", "Silent"]


def spec_check_h3_prompt(md_path, log=print):
    """对 H3 提示词 MD 执行规范自检。

    返回 (report_lines, n_error, n_warn)。report_lines 为报告文本行列表。
    """
    meta, shots = parse_h3_prompt_md(md_path)
    report = []
    n_error = 0
    n_warn = 0

    def add(level, msg):
        nonlocal n_error, n_warn
        if level == "错误":
            n_error += 1
        elif level == "警告":
            n_warn += 1
        report.append(f"[{level}] {msg}")

    report.append(f"# 规范自检报告 - {Path(md_path).name}")
    report.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append(f"镜头数: {len(shots)}")
    report.append("")

    if not shots:
        add("错误", "未解析到任何镜头（需 ## A01 九分节格式）")
        report.append("")
        report.append(f"=== 检查完成: {n_error} 错误, {n_warn} 警告 ===")
        return report, n_error, n_warn

    # 缺失分节检查
    for shot in shots:
        missing = [s for s in H3_SECTIONS if not shot.get(s)]
        if missing:
            add("错误", f"{shot['id']} 缺失分节: {', '.join(missing)}")

    # 逐镜检查
    for shot in shots:
        sid = shot["id"]
        full = shot.get("完整提示词", "")

        # 1. 指代不明（全部九个分节）
        for pat in VAGUE_REF_PATTERNS:
            for m in re.finditer(pat, full):
                line_no = full[:m.start()].count("\n") + 1
                lines = full.splitlines()
                ctx = lines[line_no - 1].strip()[:60] if line_no <= len(lines) else ""
                add("错误", f"{sid} 指代不明「{m.group()}」: {ctx}")

        # 2. 画风锚点
        style = shot.get("整体风格", "")
        if style and not any(k in style for k in STYLE_ANCHOR_KEYS):
            add("警告", f"{sid} 【整体风格】缺少画风锚点（如 D4rkL1nes 风格）")

        # 3. 画风冲突（仅查【整体风格】；"非实拍"等否定表述不算冲突）
        for k in STYLE_CONFLICT_KEYS:
            for mm in re.finditer(re.escape(k), style):
                prefix = style[max(0, mm.start() - 3):mm.start()]
                if any(c in prefix for c in "非不无"):
                    continue
                add("警告", f"{sid} 【整体风格】出现画风冲突词「{k}」（应为2D赛璐璐，非实拍/3D）")
                break

        # 4. 非叙事音乐规范
        sound = shot.get("声音", "")
        if "非叙事音乐" in sound and not any(k in sound for k in MUSIC_RULE_KEYS):
            add("警告", f"{sid} 【声音】非叙事音乐未按规范写「无背景音乐，禁止任何配乐/音乐」")

        # 5. 对白格式（语言标签 + 性别标签）
        for dt in re.findall(r"<d>(.*?)</d>", full, re.DOTALL):
            if not any(f"[{lt}]" in dt for lt in LANG_TAGS):
                add("错误", f"{sid} <d> 台词缺少语言标签: {dt.strip()[:40]}")
            if "[男]" not in dt and "[女]" not in dt and "[群杂]" not in dt:
                add("错误", f"{sid} <d> 台词缺少性别标签: {dt.strip()[:40]}")

        # 6. 时长一致性（输出规格 vs 两级时间轴）
        spec_dur = shot.get("时长", "")
        axis_dur, _ = _h3_parse_time_axis(shot.get("两级时间轴", ""))
        try:
            dur_mismatch = (spec_dur and axis_dur
                            and abs(float(spec_dur) - float(axis_dur)) > 0.01)
        except ValueError:
            dur_mismatch = False
        if dur_mismatch:
            add("警告", f"{sid} 输出规格时长({spec_dur}s)与时间轴时长({axis_dur}s)不一致")

        # 7. 参考图槽位写法
        refs = shot.get("参考图约束", "")
        if refs and "<图片" not in refs:
            add("警告", f"{sid} 【参考图约束】未使用 <图片N> 槽位写法")

    report.append("")
    report.append(f"=== 检查完成: {n_error} 错误, {n_warn} 警告 ===")
    return report, n_error, n_warn


def save_spec_report(report_lines, output_path):
    """将自检报告保存为 .md 文件。"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))


def process_single(md_path, output_path, log=print):
    """处理单个 MD 文件，返回 True/False。自动检测格式。"""
    if not Path(md_path).exists():
        log(f"ERROR: file not found: {md_path}")
        return False

    fmt = detect_format(md_path)
    if fmt == "h3_prompt":
        return process_h3_prompt_single(md_path, output_path, log=log)
    if fmt == "storyboard_v6":
        return process_storyboard_v6_single(md_path, output_path, log=log)
    if fmt == "storyboard":
        return _process_storyboard_single(md_path, output_path, log=log)
    log(f"ERROR: 无法识别 MD 格式: {md_path}")
    return False


def process_storyboard_v6_single(md_path, output_path, log=print):
    """处理单个 v6 分镜需求 MD 文件，返回 True/False。"""
    global_info, shots = parse_storyboard_v6(md_path)
    if not shots:
        log(f"ERROR: no shots parsed from {md_path}, check #### 镜头N format")
        return False

    log(f"  解析: {len(shots)} shots (v6分镜需求格式)")
    if global_info:
        log(f"  总镜头数: {global_info.get('总镜头数', '?')}, "
            f"总时长: {global_info.get('总时长', '?')}, "
            f"风格基调: {global_info.get('风格基调', '?')}")

    n_shots, n_assets = generate_storyboard_v6_excel(global_info, shots, output_path)
    log(f"  生成: {output_path} ({n_shots} shots)")
    return True


def _process_storyboard_single(md_path, output_path, log=print):
    """处理单个分镜头需求 MD 文件，返回 True/False。"""
    global_info, shots = parse_md(md_path)
    if not shots:
        log(f"ERROR: no shots parsed from {md_path}, check MD format")
        return False

    log(f"  解析: {len(shots)} shots (分镜头需求格式)")
    if global_info:
        log(f"  画幅: {global_info.get('画幅', '?')}, "
            f"总时长: {global_info.get('总时长', '?')}, "
            f"模式: {global_info.get('提示词模式', '?')}")

    n_shots, n_assets = generate_excel(global_info, shots, output_path)
    log(f"  生成: {output_path} ({n_shots} shots, {n_assets} assets)")
    return True


def process_h3_prompt_single(md_path, output_path, log=print):
    """处理单个 H3 提示词 MD 文件，返回 True/False。"""
    if not Path(md_path).exists():
        log(f"ERROR: file not found: {md_path}")
        return False

    meta, shots = parse_h3_prompt_md(md_path)
    if not shots:
        log(f"ERROR: no shots parsed from {md_path}, check ## A01 format")
        return False

    log(f"  解析: {len(shots)} shots (H3提示词格式)")
    if meta:
        log(f"  模式: {meta.get('模式', '?')}, "
            f"画风: {meta.get('画风', '?')}, "
            f"画幅: {meta.get('画幅', '?')}")

    n_shots, n_assets = generate_h3_prompt_excel(meta, shots, output_path)
    log(f"  生成: {output_path} ({n_shots} shots, {n_assets} assets)")
    return True


def process_batch(md_files, output_dir, log=print):
    """批量处理多个 MD 文件。每个 MD 生成同名 .xlsx 到 output_dir。"""
    os.makedirs(output_dir, exist_ok=True)
    success, fail = 0, 0

    for md_path in md_files:
        stem = Path(md_path).stem
        out_name = (stem.replace("分镜头需求", "提示词审阅表")
                        .replace("H3提示词", "提示词审阅表")) + ".xlsx"
        out_path = os.path.join(output_dir, out_name)
        log(f"\n[{success + fail + 1}/{len(md_files)}] {Path(md_path).name}")

        if process_single(md_path, out_path, log=log):
            success += 1
        else:
            fail += 1

    log(f"\n批量完成: {success} 成功, {fail} 失败, 共 {len(md_files)} 文件")
    return success, fail


def main():
    parser = argparse.ArgumentParser(
        description="分镜头需求 MD → 提示词审阅表 Excel"
    )
    parser.add_argument("files", nargs="+", help="MD 文件路径（一个或多个）")
    parser.add_argument("-o", "--output", help="输出目录（批量模式）或输出文件路径（单文件模式）")
    args = parser.parse_args()

    md_files = args.files

    # 单文件模式: python shot_md_to_excel.py input.md output.xlsx
    if len(md_files) == 1 and args.output and not os.path.isdir(args.output):
        if not args.output.endswith("/"):
            process_single(md_files[0], args.output)
            return
        # output 是目录，走批量
        output_dir = args.output
    elif len(md_files) == 1 and not args.output:
        # 没有指定输出，同名 .xlsx
        stem = Path(md_files[0]).stem
        out_name = (stem.replace("分镜头需求", "提示词审阅表")
                        .replace("H3提示词", "提示词审阅表")) + ".xlsx"
        process_single(md_files[0], out_name)
        return
    else:
        # 批量模式
        output_dir = args.output or os.path.dirname(md_files[0]) or "."

    process_batch(md_files, output_dir)


if __name__ == "__main__":
    main()
